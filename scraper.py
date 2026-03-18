import requests
from bs4 import BeautifulSoup
import pandas as pd

def scrape_quotes():
    base_url = 'https://quotes.toscrape.com'
    url = f'{base_url}/page/1/'
    quotes_data = []

    while url:
        print(f'Scraping {url}...')
        response = requests.get(url)
        soup = BeautifulSoup(response.text, 'html.parser')
        
        quotes = soup.find_all('div', class_='quote')
        for quote in quotes:
            text = quote.find('span', class_='text').get_text(strip=True)
            author = quote.find('small', class_='author').get_text(strip=True)
            tags = [tag.get_text(strip=True) for tag in quote.find_all('a', class_='tag')]
            
            quotes_data.append({
                'text': text,
                'author': author,
                'tags': ', '.join(tags)
            })
            
        next_btn = soup.find('li', class_='next')
        if next_btn:
            next_page_str = next_btn.find('a')['href']
            url = f'{base_url}{next_page_str}'
        else:
            url = None

    # Save to Excel using Pandas and Openpyxl
    excel_filename = 'quotes.xlsx'
    df = pd.DataFrame(quotes_data)
    
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Quotes Data')
        worksheet = writer.sheets['Quotes Data']
        
        # Import styling tools from openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        # Define the border style we want for all cells
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
                             
        # Define header background color (Dark Blue)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        # 1. Format Headers (Bold, White Text, Size 12, Centered, Dark Background)
        for cell in worksheet[1]: # Row 1 is the header
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
        # Increase the height of the header row so it stands out even more
        worksheet.row_dimensions[1].height = 25
            
        # 2. Format Data Rows (Text Wrapping and Top alignment)
        for row in worksheet.iter_rows(min_row=2, max_col=worksheet.max_column, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                
        # 3. Add borders to EVERY cell in the populated area
        for row in worksheet.iter_rows(min_row=1, max_col=worksheet.max_column, max_row=worksheet.max_row):
            for cell in row:
                cell.border = thin_border
        
        # 4. Adjust Column Widths 
        # (We set specific reasonable widths so that text wrapping actually kicks in for long quotes)
        column_widths = {
            'A': 70, # Text column (widest)
            'B': 20, # Author column
            'C': 30  # Tags column
        }
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
            
        # 5. Freeze the top row so headers stay visible while scrolling!
        worksheet.freeze_panes = 'A2'
            
    print(f'Successfully scraped {len(quotes_data)} quotes to {excel_filename}')

if __name__ == '__main__':
    scrape_quotes()
